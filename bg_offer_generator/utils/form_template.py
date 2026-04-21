"""
Form Template Generator

Creates an Excel workbook that trainee engineers can fill offline with
client-specific details. The filled template can then be uploaded back
into the Streamlit app to auto-populate the offer.

Sheets:
  1. Instructions — how to use
  2. Cover — quote ref, client name, location, contact
  3. Technical — feed parameters, technical specs, utilities
  4. Pricing — two-option pricing, payment terms, timeline
  5. Scope_Matrix — scope split between B&G and Client
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bg_offer_generator.utils.default_data import default_offer_data


BRAND_RED = "C7203E"
BRAND_PINK = "E91E63"
LIGHT_TINT = "FDECEF"


def _header_fill():
    return PatternFill(start_color=BRAND_RED, end_color=BRAND_RED, fill_type="solid")


def _section_fill():
    return PatternFill(start_color=BRAND_PINK, end_color=BRAND_PINK, fill_type="solid")


def _alt_fill():
    return PatternFill(start_color=LIGHT_TINT, end_color=LIGHT_TINT, fill_type="solid")


def _border():
    side = Side(border_style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = _header_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def _set_section_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = _section_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _border()


def _set_row(ws, row, values, alt=False):
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = _border()
        if alt:
            c.fill = _alt_fill()


def generate_form_template_xlsx() -> bytes:
    """Create and return the form template as bytes."""
    wb = Workbook()
    defaults = default_offer_data()

    # ---- Instructions ----
    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions['A'].width = 100

    ws['A1'] = "B&G Engineering Offer Form Template"
    ws['A1'].font = Font(bold=True, size=16, color=BRAND_RED)

    instructions = [
        "",
        "HOW TO USE THIS FORM",
        "",
        "1. Fill in the data across all sheets below (Cover, Technical, Pricing, Scope_Matrix).",
        "2. Yellow highlighted cells are REQUIRED. Others have sensible defaults you can leave as-is.",
        "3. All numbers should be plain numbers (no commas, no units) unless text is expected.",
        "4. Save this file.",
        "5. Go to the Streamlit offer generator app and upload this file on the 'Bulk Upload' tab.",
        "6. Review / edit on-screen, then click 'Generate DOCX'.",
        "",
        "DATA SOURCES",
        "",
        "For technical specs, utilities, equipment ratings: use the process design calculations from",
        "the bg_process_design app. Specifically from the 'Full Project JSON' download:",
        "  - plant_overview           → capacity, scheme",
        "  - stripper.results         → feed/distillate/bottoms kg/h, steam consumption",
        "  - mee.results              → effects, evaporation, concentrate, steam economy",
        "  - atfd.results             → product kg/h, moisture",
        "  - plant_wide.total_utilities → steam, power, cooling water totals",
        "  - plant_wide.economics     → operating hours, days, steam cost",
        "",
        "For commercial terms (pricing, payment, delivery): fill in based on project-specific commercials.",
        "",
        "DEFAULTS",
        "",
        "This template is pre-filled with the 100 KLD template data from a prior similar project.",
        "Adjust any numbers that differ for your specific client.",
    ]
    for i, line in enumerate(instructions, 2):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip():
            c.font = Font(bold=True, size=12, color=BRAND_RED)

    # ---- Cover sheet ----
    ws = wb.create_sheet("Cover")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    _set_header(ws.cell(row=1, column=1), "Field")
    _set_header(ws.cell(row=1, column=2), "Value")

    cov = defaults["cover"]
    cover_rows = [
        ("Quote Reference", cov["quote_ref"]),
        ("Quotation Date (YYYY-MM-DD)", cov["quote_date"]),
        ("Submitted to (Client)", cov["submitted_to"]),
        ("Location", cov["location"]),
        ("Prepared By", cov["prepared_by"]),
        ("Contact Details", cov["contact_details"]),
        ("E-mail", cov["email"]),
        ("Kind Attn (Client Contact)", cov["kind_attn"]),
        ("Subject Line", cov["subject"]),
        ("Discussion Date", cov["discussion_date"]),
        ("Capacity (KLD)", cov["capacity_kld"]),
    ]
    for i, (k, v) in enumerate(cover_rows, 2):
        _set_row(ws, i, [k, v], alt=(i % 2 == 0))

    # ---- Technical sheet ----
    ws = wb.create_sheet("Technical")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

    row = 1
    _set_header(ws.cell(row=row, column=1), "Parameter")
    _set_header(ws.cell(row=row, column=2), "UOM")
    _set_header(ws.cell(row=row, column=3), "Value")
    row += 1

    # Feed parameters section
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "FEED PARAMETERS")
    row += 1

    fp = defaults["feed_parameters"]
    feed_rows = [
        ("Feed / Capacity", "KLD", fp["capacity_kld"]),
        ("Feed pH", "-", fp["feed_ph"]),
        ("Total COD", "PPM", fp["total_cod_ppm"]),
        ("Volatile Organic Solvents", "PPM", fp["volatile_organic_solvents_ppm"]),
        ("Total Solids", "% w/w", fp["total_solids_pct"]),
        ("Suspended Solids", "PPM", fp["suspended_solids_ppm"]),
        ("Feed Temperature", "°C", fp["feed_temp_c"]),
        ("Total Hardness", "PPM", fp["total_hardness_ppm"]),
        ("Silica", "PPM", fp["silica_ppm"]),
        ("Free Chloride", "PPM", fp["free_chloride_ppm"]),
        ("Feed Nature", "-", fp["feed_nature"]),
    ]
    for i, rdata in enumerate(feed_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # Stripper specs
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "STRIPPER TECHNICAL SPECS")
    row += 1
    ts = defaults["technical_specs"]
    stripper_rows = [
        ("Stripper Type", "-", ts["stripper"]["type"]),
        ("Feed Inlet", "Kg/h", ts["stripper"]["feed_kgh"]),
        ("Top Distillate Out", "Kg/h", ts["stripper"]["distillate_kgh"]),
        ("Distillate Composition", "-", ts["stripper"]["distillate_composition"]),
        ("Stripper Bottom Out", "Kg/h", ts["stripper"]["bottoms_kgh"]),
    ]
    for i, rdata in enumerate(stripper_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # MEE specs
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "MEE TECHNICAL SPECS")
    row += 1
    mee_rows = [
        ("MEE Type (e.g. '4-Effect MEE')", "-", ts["mee"]["type"]),
        ("Configuration", "-", ts["mee"]["configuration"]),
        ("Feed Inlet (Stripper Bottom)", "Kg/h", ts["mee"]["feed_kgh"]),
        ("Feed Solids", "%", ts["mee"]["feed_solids_pct"]),
        ("Water Evaporation Rate", "Kg/h", ts["mee"]["evaporation_kgh"]),
        ("MEE Concentrate Out", "Kg/h", ts["mee"]["concentrate_kgh"]),
        ("Concentrate Solids", "%", ts["mee"]["concentrate_solids_pct"]),
    ]
    for i, rdata in enumerate(mee_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # ATFD specs
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "ATFD TECHNICAL SPECS")
    row += 1
    atfd_rows = [
        ("ATFD Type", "-", ts["atfd"]["type"]),
        ("Feed Inlet (MEE Concentrate)", "Kg/h", ts["atfd"]["feed_kgh"]),
        ("Feed Solids", "%", ts["atfd"]["feed_solids_pct"]),
        ("Water Evaporation Rate", "Kg/h", ts["atfd"]["evaporation_kgh"]),
        ("ATFD Product Out", "Kg/h", ts["atfd"]["product_kgh"]),
        ("Moisture in Product", "%", ts["atfd"]["product_moisture_pct"]),
    ]
    for i, rdata in enumerate(atfd_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # Utilities
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "UTILITIES")
    row += 1
    ut = defaults["utilities"]
    util_rows = [
        ("Stripper Steam Parameters", "-", ut["stripper_steam"]["param"]),
        ("Stripper Steam Flow", "Kg/h", ut["stripper_steam"]["value_kgh"]),
        ("MEE Steam Parameters", "-", ut["mee_steam"]["param"]),
        ("MEE Steam Flow", "Kg/h", ut["mee_steam"]["value_kgh"]),
        ("MEE Steam Economy", "Kg/Kg", ut["mee_steam"]["steam_economy"]),
        ("ATFD Steam Parameters", "-", ut["atfd_steam"]["param"]),
        ("ATFD Steam Flow", "Kg/h", ut["atfd_steam"]["value_kgh"]),
        ("Power Consumption", "kWh", ut["power_consumption_kwh"]),
        ("Power Installed (incl. standby)", "kW", ut["power_installed_kw"]),
        ("Cooling Water Flow", "m³/h", ut["cooling_water_m3h"]),
        ("Cooling Water Temperatures", "-", ut["cooling_water_temps"]),
        ("Compressed Air Flow", "Nm³/h", ut["compressed_air_nm3h"]),
        ("Compressed Air Pressure", "-", ut["compressed_air_pressure"]),
    ]
    for i, rdata in enumerate(util_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # Economics
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_section_header(ws.cell(row=row, column=1), "ECONOMICS / OPEX")
    row += 1
    econ = defaults["economics"]
    econ_rows = [
        ("Conventional Steam Consumption", "Kg/h", econ["conventional_steam_kgh"]),
        ("ECOX-ZLD Steam Consumption", "Kg/h", econ["ecox_steam_kgh"]),
        ("Steam Reduction %", "%", econ["steam_reduction_pct"]),
        ("Conventional Annual Steam", "tons/yr", econ["conventional_annual_steam_tons"]),
        ("ECOX Annual Steam", "tons/yr", econ["ecox_annual_steam_tons"]),
        ("Annual Steam Savings", "tons/yr", econ["annual_steam_savings_tons"]),
        ("Conventional Annual OPEX", "Cr/yr", econ["conventional_annual_cost_cr"]),
        ("ECOX Annual OPEX", "Cr/yr", econ["ecox_annual_cost_cr"]),
        ("Annual Savings", "Lakhs/yr", econ["annual_savings_lakhs"]),
        ("Operating Hours/Day", "-", econ["operating_hours_day"]),
        ("Operating Days/Year", "-", econ["operating_days_year"]),
        ("Steam Cost", "INR/Kg", econ["steam_cost_inr_kg"]),
    ]
    for i, rdata in enumerate(econ_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # ---- Pricing sheet ----
    ws = wb.create_sheet("Pricing")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40

    row = 1
    _set_header(ws.cell(row=row, column=1), "Field")
    _set_header(ws.cell(row=row, column=2), "Value")
    row += 1

    pr = defaults["pricing"]
    pricing_rows = [
        ("Option 1 MOC", pr["option1_moc"]),
        ("Option 2 MOC", pr["option2_moc"]),
        ("Option 1 — Equipment Price (Cr)", pr["option1_equipment_price_cr"]),
        ("Option 2 — Equipment Price (Cr)", pr["option2_equipment_price_cr"]),
        ("Option 1 — Install & Commissioning (Lakhs)", pr["option1_install_lakhs"]),
        ("Option 2 — Install & Commissioning (Lakhs)", pr["option2_install_lakhs"]),
        ("Option 1 — Total (Cr)", pr["option1_total_cr"]),
        ("Option 2 — Total (Cr)", pr["option2_total_cr"]),
        ("Location DAP", pr["location_dap"]),
        ("Price Validity (Days)", pr["price_validity_days"]),
        ("Payment Term 1", pr["payment_terms"][0]),
        ("Payment Term 2", pr["payment_terms"][1]),
        ("Payment Term 3", pr["payment_terms"][2]),
        ("Delivery — Option 1", pr["delivery_timeline"]["supply_option1"]),
        ("Delivery — Option 2", pr["delivery_timeline"]["supply_option2"]),
        ("Installation Timeline", pr["delivery_timeline"]["installation"]),
        ("Commissioning Timeline", pr["delivery_timeline"]["commissioning"]),
    ]
    for i, rdata in enumerate(pricing_rows):
        _set_row(ws, row, rdata, alt=(i % 2 == 1))
        row += 1

    # ---- Scope Matrix ----
    ws = wb.create_sheet("Scope_Matrix")
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    _set_header(ws.cell(row=1, column=1), "S.No")
    _set_header(ws.cell(row=1, column=2), "Description of Work")
    _set_header(ws.cell(row=1, column=3), "B&G")
    _set_header(ws.cell(row=1, column=4), "Client")

    for i, item in enumerate(defaults["scope_matrix"], 2):
        _set_row(ws, i, [
            i - 1,
            item["description"],
            "Yes" if item["bg"] else "No",
            "Yes" if item["client"] else "No",
        ], alt=(i % 2 == 0))

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
