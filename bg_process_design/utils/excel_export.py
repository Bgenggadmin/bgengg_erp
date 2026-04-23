"""
Review Excel export for B&G Process Design projects.

Produces a 5-sheet branded workbook with all design outputs laid out for manager
review. All static values (not formulas) — purely a read-only report view of
what the ERP has calculated.

Sheets:
  1. Project Summary   — project facts + KPIs + economics headlines
  2. Stripper          — inputs + results + equipment
  3. MEE               — inputs + effects + utilities + equipment
  4. ATFD              — inputs + results + equipment
  5. Plant-Wide        — utilities + pumps + feed characterization + economics

Usage:
    from bg_process_design.utils.excel_export import build_review_workbook
    xlsx_bytes = build_review_workbook(project_data)
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------
# Brand palette (as hex strings, no #)
# -----------------------------------------------------------------
RED         = 'C7203E'
PINK        = 'E91E63'
CHARCOAL    = '1E2A38'
MID_GREY    = '4A5568'
LIGHT_GREY  = 'E2E8F0'
BG_LIGHT    = 'F7F9FC'
WHITE       = 'FFFFFF'
HEADER_BG   = 'F5E8EA'  # very light pink tint for section headers
PURPLE      = '8E24AA'
BLUE        = '0284C7'
GREEN       = '059669'

# -----------------------------------------------------------------
# Shared styles
# -----------------------------------------------------------------
FONT_NAME = 'Calibri'

TITLE_FONT      = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
SUBTITLE_FONT   = Font(name=FONT_NAME, size=11, italic=True, color=WHITE)
SECTION_FONT    = Font(name=FONT_NAME, size=12, bold=True, color=RED)
LABEL_FONT      = Font(name=FONT_NAME, size=10, color=MID_GREY)
VALUE_FONT      = Font(name=FONT_NAME, size=10, bold=True, color=CHARCOAL)
HEADER_FONT     = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
KPI_LABEL_FONT  = Font(name=FONT_NAME, size=9, bold=True, color=MID_GREY)
KPI_VALUE_FONT  = Font(name=FONT_NAME, size=22, bold=True, color=RED)
KPI_UNIT_FONT   = Font(name=FONT_NAME, size=9, color=MID_GREY)
NOTE_FONT       = Font(name=FONT_NAME, size=9, italic=True, color=MID_GREY)

FILL_TITLE      = PatternFill('solid', start_color=CHARCOAL, end_color=CHARCOAL)
FILL_HEADER     = PatternFill('solid', start_color=RED, end_color=RED)
FILL_SECTION    = PatternFill('solid', start_color=HEADER_BG, end_color=HEADER_BG)
FILL_ALT_ROW    = PatternFill('solid', start_color=BG_LIGHT, end_color=BG_LIGHT)
FILL_KPI_CARD   = PatternFill('solid', start_color=BG_LIGHT, end_color=BG_LIGHT)

ALIGN_LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT     = Alignment(horizontal='right',  vertical='center', wrap_text=True)

THIN_BORDER     = Border(
    left=Side(style='thin', color=LIGHT_GREY),
    right=Side(style='thin', color=LIGHT_GREY),
    top=Side(style='thin', color=LIGHT_GREY),
    bottom=Side(style='thin', color=LIGHT_GREY),
)
BOTTOM_BORDER = Border(bottom=Side(style='thin', color=LIGHT_GREY))


# =================================================================
# Helpers
# =================================================================
def _fmt(v, places=1, suffix=''):
    """Format a numeric value for display, handling None/0 gracefully."""
    if v is None:
        return '—'
    try:
        if isinstance(v, int) and places == 0:
            return f"{v:,}{suffix}"
        v = float(v)
        if abs(v) >= 1000 and places <= 1:
            return f"{v:,.0f}{suffix}"
        return f"{v:,.{places}f}{suffix}"
    except (ValueError, TypeError):
        return str(v)


def _set_col_widths(ws, widths: dict) -> None:
    """widths = {'A': 18, 'B': 25, ...}"""
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _title_bar(ws, row: int, title: str, subtitle: str = '', span_cols: int = 6) -> int:
    """Draw a charcoal title bar starting at `row`. Returns next free row."""
    # Title cell
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = FILL_TITLE
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 32

    # Subtitle row
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle)
        ws.cell(row=row + 1, column=1).font = SUBTITLE_FONT
        ws.cell(row=row + 1, column=1).fill = FILL_TITLE
        ws.cell(row=row + 1, column=1).alignment = Alignment(
            horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row + 1, start_column=1,
                       end_row=row + 1, end_column=span_cols)
        ws.row_dimensions[row + 1].height = 20
        return row + 3  # blank row after
    return row + 2


def _section_header(ws, row: int, text: str, span_cols: int = 6) -> int:
    """Draw a pink-tinted section header bar. Returns next free row."""
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.cell(row=row, column=1).alignment = Alignment(
        horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 22
    return row + 1


def _data_table(ws, start_row: int, headers: list, rows: list,
                col_widths: Optional[list] = None) -> int:
    """
    Draw a data table with pink header row and alternating background.
    headers: list of column titles
    rows:    list of tuples (one per data row)
    Returns next free row after the table.
    """
    # Header row
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 22

    # Data rows
    for r_offset, row_vals in enumerate(rows):
        r = start_row + 1 + r_offset
        fill = FILL_ALT_ROW if r_offset % 2 == 1 else None
        for i, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = VALUE_FONT if i > 1 else Font(name=FONT_NAME, size=10, color=CHARCOAL)
            cell.alignment = ALIGN_RIGHT if i > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[r].height = 18

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    return start_row + 1 + len(rows) + 1  # +1 blank row


def _key_value_block(ws, start_row: int, items: list,
                     label_col: str = 'A', value_col: str = 'B') -> int:
    """
    Draw a two-column label/value list (no header row).
    items: list of (label, value) tuples.
    Returns next free row.
    """
    for i, (label, value) in enumerate(items):
        r = start_row + i
        c_label = ws.cell(row=r, column=1, value=label)
        c_value = ws.cell(row=r, column=2, value=value)
        c_label.font = LABEL_FONT
        c_label.alignment = ALIGN_LEFT
        c_label.border = BOTTOM_BORDER
        c_value.font = VALUE_FONT
        c_value.alignment = ALIGN_RIGHT
        c_value.border = BOTTOM_BORDER
        ws.row_dimensions[r].height = 18

    return start_row + len(items) + 1  # blank after


def _kpi_row(ws, row: int, kpis: list) -> int:
    """
    Draw a horizontal row of KPI cards. Each KPI spans 2 columns.
    kpis: list of (label, value, unit) tuples — max 4 recommended.
    Returns next free row (card takes 3 rows: label, value, unit).
    """
    for i, (label, value, unit) in enumerate(kpis):
        c_start = 1 + i * 2
        # Label
        ws.cell(row=row, column=c_start, value=label)
        ws.cell(row=row, column=c_start).font = KPI_LABEL_FONT
        ws.cell(row=row, column=c_start).fill = FILL_KPI_CARD
        ws.cell(row=row, column=c_start).alignment = Alignment(
            horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row, start_column=c_start,
                       end_row=row, end_column=c_start + 1)

        # Value
        ws.cell(row=row + 1, column=c_start, value=value)
        ws.cell(row=row + 1, column=c_start).font = KPI_VALUE_FONT
        ws.cell(row=row + 1, column=c_start).fill = FILL_KPI_CARD
        ws.cell(row=row + 1, column=c_start).alignment = Alignment(
            horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row + 1, start_column=c_start,
                       end_row=row + 1, end_column=c_start + 1)

        # Unit
        ws.cell(row=row + 2, column=c_start, value=unit)
        ws.cell(row=row + 2, column=c_start).font = KPI_UNIT_FONT
        ws.cell(row=row + 2, column=c_start).fill = FILL_KPI_CARD
        ws.cell(row=row + 2, column=c_start).alignment = Alignment(
            horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row + 2, start_column=c_start,
                       end_row=row + 2, end_column=c_start + 1)

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 32
    ws.row_dimensions[row + 2].height = 16

    return row + 4  # 3 KPI rows + 1 blank


# =================================================================
# Sheet builders
# =================================================================
def _build_summary_sheet(ws, proj, po, pw, e, has_strip, has_mee, has_atfd) -> None:
    _set_col_widths(ws, {
        'A': 26, 'B': 26, 'C': 26, 'D': 26, 'E': 26, 'F': 14,
    })

    scheme = proj.get('scheme') or 'Stripper + MEE + ATFD'
    row = _title_bar(ws, 1,
                     f"B&G ENGINEERING   •   {proj.get('project_code', '').upper() or 'PROJECT'}",
                     f"{scheme}   •   Responsible towards water",
                     span_cols=6)

    # ---- Project facts ----
    row = _section_header(ws, row, '  PROJECT DETAILS', span_cols=6)
    facts = [
        ('Project Code',        proj.get('project_code') or '—'),
        ('Project Name',        proj.get('project_name') or '—'),
        ('Client ID',           proj.get('client_id') or '—'),
        ('Plant Location',      proj.get('plant_location') or '—'),
        ('Capacity',            f"{proj.get('capacity_kld', '—')} KLD"),
        ('Process Scheme',      scheme),
        ('Designed By',         proj.get('designed_by') or '—'),
        ('Status',              (proj.get('status') or '—').upper()),
        ('Notes',               proj.get('notes') or '—'),
        ('Report Generated',    datetime.now().strftime('%d %b %Y, %H:%M')),
    ]
    row = _key_value_block(ws, row, facts)
    row += 1

    # ---- Design status ----
    row = _section_header(ws, row, '  DESIGN COMPLETENESS', span_cols=6)
    status_items = [
        ('Stripper',  '✅ Designed' if has_strip else '⚪ Not designed'),
        ('MEE',       '✅ Designed' if has_mee else '⚪ Not designed'),
        ('ATFD',      '✅ Designed' if has_atfd else '⚪ Not designed'),
    ]
    row = _key_value_block(ws, row, status_items)
    row += 1

    # ---- KPIs ----
    row = _section_header(ws, row, '  KEY PERFORMANCE INDICATORS', span_cols=6)
    steam_econ = po.get('mee_steam_economy') or 0
    total_steam = (pw.get('total_utilities') or {}).get('steam_kgh') or 0
    daily_opex = (e or {}).get('total_daily_op_cost_inr') or 0
    daily_lakhs = daily_opex / 100000 if daily_opex else 0

    row = _kpi_row(ws, row, [
        ('CAPACITY',    _fmt(proj.get('capacity_kld', 0), 0),           'KLD'),
        ('STEAM ECON.', _fmt(steam_econ, 2),                            ': 1 MEE'),
        ('TOTAL STEAM', _fmt(total_steam / 1000 if total_steam else 0, 1), 'TPH'),
        ('DAILY OPEX',  f"₹ {_fmt(daily_lakhs, 2)}",                    'LAKH / DAY'),
    ])

    # ---- Economics headline ----
    if e:
        row = _section_header(ws, row, '  OPERATING ECONOMICS', span_cols=6)
        econ_items = [
            ('Cost per KL treated',       f"₹ {_fmt(e.get('cost_per_kl_inr'), 0)}"),
            ('Daily operating cost',      f"₹ {_fmt(e.get('total_daily_op_cost_inr'), 0)}   ({_fmt(daily_lakhs, 2)} lakh/day)"),
            ('Annual operating cost',     f"₹ {_fmt(e.get('annual_op_cost_inr'), 0)}   ({_fmt((e.get('annual_op_cost_inr') or 0) / 10000000, 2)} crore/yr)"),
            ('Operating days / year',     f"{e.get('operating_days_year', '—')}"),
            ('Operating hours / day',     f"{e.get('operating_hours_day', '—')}"),
            ('  — Steam cost / day',      f"₹ {_fmt(e.get('daily_steam_cost_inr'), 0)}"),
            ('  — Power cost / day',      f"₹ {_fmt(e.get('daily_power_cost_inr'), 0)}"),
            ('  — Cooling water / day',   f"₹ {_fmt(e.get('daily_cw_cost_inr'), 0)}"),
        ]
        row = _key_value_block(ws, row, econ_items)
        row += 1

    # ---- Plant overview ----
    row = _section_header(ws, row, '  PLANT OVERVIEW', span_cols=6)
    po_items = [
        ('Process scheme',            po.get('process_scheme') or '—'),
        ('Process flow',              po.get('process_flow') or '—'),
        ('Stripper capacity',         f"{_fmt(po.get('stripper_capacity_kgh'), 0)} kg/h"),
        ('Solvent recovered',         f"{_fmt(po.get('solvent_recovered_kgh'), 0)} kg/h"),
        ('MEE capacity',              f"{_fmt(po.get('mee_capacity_kgh'), 0)} kg/h"),
        ('MEE effects',               f"{po.get('mee_n_effects') or '—'}"),
        ('MEE steam economy',         f"{_fmt(po.get('mee_steam_economy'), 2)} : 1"),
        ('MEE total evaporation',     f"{_fmt(po.get('mee_total_evaporation_kgh'), 0)} kg/h"),
        ('ATFD dry product',          f"{_fmt(po.get('atfd_dry_product_kgh'), 0)} kg/h"),
        ('ATFD product solids',       f"{_fmt(po.get('atfd_product_ts_pct'), 1)} %"),
    ]
    _key_value_block(ws, row, po_items)


def _build_stripper_sheet(ws, strip_i, strip_r) -> None:
    _set_col_widths(ws, {'A': 32, 'B': 24, 'C': 24, 'D': 24, 'E': 18})

    row = _title_bar(ws, 1, 'STAGE 1 — STRIPPER COLUMN',
                     'Steam-stripped distillation for solvent recovery',
                     span_cols=5)

    if not strip_r or not strip_r.get('feed_kgh'):
        _section_header(ws, row, '  No Stripper design saved for this project', span_cols=5)
        return

    # ---- Design inputs ----
    row = _section_header(ws, row, '  DESIGN INPUTS', span_cols=5)
    input_items = [
        ('Feed rate',                f"{_fmt(strip_i.get('feed_rate_kgh'), 0)} kg/h"),
        ('Feed temperature',         f"{_fmt(strip_i.get('feed_temp_c'), 1)} °C"),
        ('Solvent fraction',         f"{_fmt((strip_i.get('solids_frac') or 0) * 100, 1)} % w/w"),
        ('Solvent mix',              ', '.join(f"{k}: {v*100:.0f}%" for k, v in (strip_i.get('solvent_mix') or {}).items()) or '—'),
        ('Water fraction',           f"{_fmt((strip_i.get('water_frac') or 0) * 100, 1)} %"),
        ('Required recovery',        f"{_fmt((strip_i.get('solvent_recovery') or 0) * 100, 0)} %"),
        ('Number of trays',          f"{strip_i.get('no_of_trays') or '—'}"),
        ('Tray spacing',             f"{_fmt((strip_i.get('tray_spacing_m') or 0) * 1000, 0)} mm"),
        ('Reflux ratio',             f"{_fmt(strip_i.get('reflux_ratio'), 2)}"),
        ('Steam pressure',           f"{_fmt(strip_i.get('steam_pressure_bar'), 1)} bar"),
        ('Liquid density',           f"{_fmt(strip_i.get('liquid_density_kgm3'), 0)} kg/m³"),
        ('CW inlet / outlet',        f"{_fmt(strip_i.get('cw_in_c'), 0)} °C / {_fmt(strip_i.get('cw_out_c'), 0)} °C"),
    ]
    row = _key_value_block(ws, row, input_items)
    row += 1

    # ---- Mass balance ----
    row = _section_header(ws, row, '  MASS BALANCE', span_cols=5)
    mb = [
        ('Feed rate',                f"{_fmt(strip_r.get('feed_kgh'), 0)} kg/h"),
        ('Solvent input',            f"{_fmt(strip_r.get('solvent_in_kgh'), 0)} kg/h"),
        ('Distillate (overhead)',    f"{_fmt(strip_r.get('distillate_kgh'), 0)} kg/h"),
        ('Bottoms (→ MEE)',           f"{_fmt(strip_r.get('bottoms_kgh'), 0)} kg/h"),
        ('Water evaporated',         f"{_fmt(strip_r.get('water_evap_kgh'), 0)} kg/h"),
        ('Reflux pump flow',         f"{_fmt(strip_r.get('reflux_pump_flow_m3h'), 2)} m³/h"),
        ('Vapor flow',               f"{_fmt(strip_r.get('vapor_flow_kgh'), 0)} kg/h"),
        ('Liquid flow',              f"{_fmt(strip_r.get('liquid_flow_kgh'), 0)} kg/h"),
    ]
    row = _key_value_block(ws, row, mb)
    row += 1

    # ---- Column sizing ----
    row = _section_header(ws, row, '  COLUMN SIZING', span_cols=5)
    cs = [
        ('Column diameter (calc)',   f"{_fmt((strip_r.get('column_dia_calc_m') or 0) * 1000, 0)} mm"),
        ('Column diameter (selected)', f"{_fmt((strip_r.get('column_dia_selected_m') or 0) * 1000, 0)} mm"),
        ('Column area',              f"{_fmt(strip_r.get('column_area_m2'), 3)} m²"),
        ('Design velocity',          f"{_fmt(strip_r.get('design_velocity_ms'), 3)} m/s"),
        ('Flooding velocity',        f"{_fmt(strip_r.get('flooding_velocity_ms'), 3)} m/s"),
        ('Number of trays',          f"{strip_r.get('no_of_trays') or '—'}"),
        ('Tray spacing',             f"{_fmt((strip_r.get('tray_spacing_m') or 0) * 1000, 0)} mm"),
        ('Weir length',              f"{_fmt(strip_r.get('weir_length_m'), 3)} m"),
        ('Weir height',              f"{_fmt(strip_r.get('weir_height_mm'), 1)} mm"),
        ('Total pressure drop',      f"{_fmt(strip_r.get('dp_total_mm'), 0)} mm H₂O   ({_fmt(strip_r.get('dp_total_bar'), 3)} bar)"),
    ]
    row = _key_value_block(ws, row, cs)
    row += 1

    # ---- Reboiler ----
    reb = strip_r.get('reboiler_tubes') or {}
    row = _section_header(ws, row, '  REBOILER', span_cols=5)
    rb = [
        ('HTA calculated',           f"{_fmt(strip_r.get('reboiler_HTA_calc'), 2)} m²"),
        ('HTA selected',             f"{_fmt(strip_r.get('reboiler_HTA_selected'), 0)} m²"),
        ('LMTD',                     f"{_fmt(strip_r.get('reboiler_lmtd'), 1)} °C"),
        ('Overall U',                f"{_fmt(strip_r.get('reboiler_U'), 0)} kcal/h·m²·°C"),
        ('Heat load',                f"{_fmt(strip_r.get('reboiler_heat_load_kcalh'), 0)} kcal/h"),
        ('Shell temperature',        f"{_fmt(strip_r.get('reboiler_shell_temp'), 1)} °C"),
        ('Tube inlet / outlet',      f"{_fmt(strip_r.get('reboiler_inlet_temp'), 1)} °C / {_fmt(strip_r.get('reboiler_outlet_temp'), 1)} °C"),
        ('Tubes',                    f"{reb.get('total_tubes', '—')} × {_fmt(reb.get('tube_od_mm'), 1)} mm OD × {_fmt(reb.get('tube_length_m'), 1)} m"),
        ('Tube passes',              f"{reb.get('n_passes', '—')}"),
        ('Shell ID (selected)',      f"{_fmt((reb.get('shell_id_selected_m') or 0) * 1000, 0)} mm"),
        ('Tube velocity',            f"{_fmt(reb.get('tube_velocity_ms'), 2)} m/s"),
    ]
    row = _key_value_block(ws, row, rb)
    row += 1

    # ---- Condenser ----
    cond = strip_r.get('condenser1_tubes') or {}
    row = _section_header(ws, row, '  CONDENSER', span_cols=5)
    cn = [
        ('HTA calculated',           f"{_fmt(strip_r.get('condenser1_HTA_calc'), 2)} m²"),
        ('HTA selected',             f"{_fmt(strip_r.get('condenser1_HTA_selected'), 0)} m²"),
        ('LMTD',                     f"{_fmt(strip_r.get('condenser1_lmtd'), 1)} °C"),
        ('Overall U',                f"{_fmt(strip_r.get('condenser1_U'), 0)} kcal/h·m²·°C"),
        ('CW flow',                  f"{_fmt(strip_r.get('cw_flow_m3h'), 2)} m³/h"),
        ('Tubes',                    f"{cond.get('total_tubes', '—')} × {_fmt(cond.get('tube_od_mm'), 1)} mm OD × {_fmt(cond.get('tube_length_m'), 1)} m"),
        ('Tube passes',              f"{cond.get('n_passes', '—')}"),
        ('Shell ID (selected)',      f"{_fmt((cond.get('shell_id_selected_m') or 0) * 1000, 0)} mm"),
    ]
    row = _key_value_block(ws, row, cn)
    row += 1

    # ---- Pumps ----
    pumps = strip_r.get('pumps') or {}
    if pumps:
        row = _section_header(ws, row, '  PUMPS', span_cols=5)
        pump_rows = []
        for pid, p in pumps.items():
            pump_rows.append((
                p.get('service', pid),
                _fmt(p.get('flow_m3h'), 2) + ' m³/h',
                _fmt(p.get('head_mlc'), 0) + ' m',
                _fmt(p.get('brake_power_kw'), 2) + ' kW',
                _fmt(p.get('motor_hp_selected'), 1) + ' HP',
            ))
        row = _data_table(ws, row,
                          headers=['Service', 'Flow', 'Head', 'Brake Power', 'Motor'],
                          rows=pump_rows,
                          col_widths=[32, 16, 12, 16, 14])


def _build_mee_sheet(ws, mee_i, mee_r) -> None:
    _set_col_widths(ws, {'A': 30, 'B': 22, 'C': 22, 'D': 22, 'E': 18, 'F': 14})

    row = _title_bar(ws, 1, 'STAGE 2 — MULTI-EFFECT EVAPORATOR',
                     f"{mee_r.get('n_effects', 4)}-effect falling-film evaporator",
                     span_cols=6)

    if not mee_r or not mee_r.get('feed_kgh'):
        _section_header(ws, row, '  No MEE design saved for this project', span_cols=6)
        return

    # ---- Inputs ----
    row = _section_header(ws, row, '  DESIGN INPUTS', span_cols=6)
    input_items = [
        ('Number of effects',        f"{mee_r.get('n_effects') or '—'}"),
        ('Feed rate',                f"{_fmt(mee_r.get('feed_kgh'), 0)} kg/h"),
        ('Feed total solids',        f"{_fmt(mee_r.get('feed_ts_pct'), 2)} %"),
        ('Outlet total solids',      f"{_fmt(mee_r.get('outlet_ts_pct'), 1)} %"),
        ('Auto BPR used',            'Yes' if mee_r.get('auto_bpr_used') else 'No'),
        ('Stripper vapor integrated', f"{_fmt(mee_r.get('stripper_vapor_integrated_kgh'), 0)} kg/h"),
    ]
    row = _key_value_block(ws, row, input_items)
    row += 1

    # ---- Results ----
    row = _section_header(ws, row, '  MASS BALANCE & PERFORMANCE', span_cols=6)
    mb = [
        ('Total evaporation',        f"{_fmt(mee_r.get('total_evap_kgh'), 0)} kg/h"),
        ('Final concentrate',        f"{_fmt(mee_r.get('final_concentrate_kgh'), 0)} kg/h"),
        ('Live steam consumption',   f"{_fmt(mee_r.get('steam_consumption_kgh'), 0)} kg/h"),
        ('Steam economy',            f"{_fmt(mee_r.get('steam_economy'), 2)} : 1"),
    ]
    row = _key_value_block(ws, row, mb)
    row += 1

    # ---- Effect-by-effect table ----
    effects = mee_r.get('effects') or []
    if effects:
        row = _section_header(ws, row, '  EFFECT-BY-EFFECT DETAILS', span_cols=6)
        final_conc = mee_r.get('final_concentrate_kgh') or 0
        total_evap = mee_r.get('total_evap_kgh') or 1

        effect_rows = []
        for i, e in enumerate(effects):
            feed = e.get('feed_kgh') or 0
            if i < len(effects) - 1:
                next_feed = effects[i + 1].get('feed_kgh') or 0
                evap = feed - next_feed
            else:
                evap = feed - final_conc
            pct = (evap / total_evap * 100) if total_evap else 0
            role = 'Steam-driven' if i == 0 else ('Vacuum' if i == len(effects) - 1 else 'Vapor-driven')
            effect_rows.append((
                f"E-0{e.get('effect_no', i+1)}",
                _fmt(feed, 0),
                _fmt(evap, 0),
                f"{_fmt(pct, 1)} %",
                _fmt(e.get('shell_temp_c'), 1),
                role,
            ))
        row = _data_table(ws, row,
                          headers=['Effect', 'Feed (kg/h)', 'Evaporated (kg/h)', '% of Total', 'Shell Temp (°C)', 'Role'],
                          rows=effect_rows,
                          col_widths=[14, 16, 18, 14, 18, 18])
        row += 1

    # ---- Utilities ----
    u = mee_r.get('utilities') or {}
    if u:
        row = _section_header(ws, row, '  UTILITIES', span_cols=6)
        ui = [
            ('Live steam',               f"{_fmt(u.get('steam_kgh'), 0)} kg/h"),
            ('Power demand',             f"{_fmt(u.get('power_kw'), 1)} kW"),
            ('Cooling water circulation', f"{_fmt(u.get('cw_m3h'), 0)} m³/h"),
            ('Cooling water makeup',     f"{_fmt(u.get('cw_makeup_m3h'), 2)} m³/h"),
        ]
        row = _key_value_block(ws, row, ui)
        row += 1

    # ---- Condenser ----
    cond = mee_r.get('condenser') or {}
    if cond:
        row = _section_header(ws, row, '  FINAL CONDENSER', span_cols=6)
        ci = [
            ('HTA calculated',           f"{_fmt(cond.get('HTA_calc_m2'), 1)} m²"),
            ('HTA selected',             f"{_fmt(cond.get('HTA_selected_m2'), 0)} m²"),
            ('Heat load',                f"{_fmt(cond.get('heat_load_kcalh'), 0)} kcal/h"),
            ('LMTD',                     f"{_fmt(cond.get('lmtd_c'), 1)} °C"),
            ('Vapor inlet',              f"{_fmt(cond.get('vapor_in_kgh'), 0)} kg/h"),
            ('CW flow',                  f"{_fmt(cond.get('cw_flow_m3h'), 0)} m³/h"),
            ('CW inlet / outlet',        f"{_fmt(cond.get('cw_in_c'), 1)} °C / {_fmt(cond.get('cw_out_c'), 1)} °C"),
        ]
        row = _key_value_block(ws, row, ci)
        row += 1

    # ---- Pumps summary ----
    pumps = mee_r.get('pumps') or {}
    if pumps:
        row = _section_header(ws, row, '  MEE PUMPS', span_cols=6)
        pump_rows = []
        for pid, p in pumps.items():
            pump_rows.append((
                p.get('service', pid),
                _fmt(p.get('flow_m3h'), 2) + ' m³/h',
                _fmt(p.get('head_mlc'), 0) + ' m',
                _fmt(p.get('brake_power_kw'), 2) + ' kW',
                _fmt(p.get('motor_hp_selected'), 1) + ' HP',
            ))
        row = _data_table(ws, row,
                          headers=['Service', 'Flow', 'Head', 'Brake Power', 'Motor'],
                          rows=pump_rows,
                          col_widths=[32, 16, 12, 16, 14])


def _build_atfd_sheet(ws, atfd_i, atfd_r) -> None:
    _set_col_widths(ws, {'A': 32, 'B': 28, 'C': 24, 'D': 22, 'E': 18})

    row = _title_bar(ws, 1, 'STAGE 3 — AGITATED THIN FILM DRYER',
                     'Final drying stage — closes the zero-discharge loop',
                     span_cols=5)

    if not atfd_r or not atfd_r.get('feed_kgh'):
        _section_header(ws, row, '  No ATFD design saved for this project', span_cols=5)
        return

    # ---- Mass balance ----
    row = _section_header(ws, row, '  MASS BALANCE', span_cols=5)
    mb = [
        ('Feed rate',                f"{_fmt(atfd_r.get('feed_kgh'), 0)} kg/h"),
        ('Feed total solids',        f"{_fmt(atfd_r.get('feed_ts_pct'), 1)} %"),
        ('Water evaporated',         f"{_fmt(atfd_r.get('water_evap_kgh'), 0)} kg/h"),
        ('Water in feed',            f"{_fmt(atfd_r.get('water_in_kgh'), 0)} kg/h"),
        ('Dry product output',       f"{_fmt(atfd_r.get('product_kgh'), 0)} kg/h"),
        ('Product total solids',     f"{_fmt(atfd_r.get('product_ts_pct'), 1)} %"),
        ('Solids production',        f"{_fmt(atfd_r.get('solids_kgh'), 2)} kg/h"),
    ]
    row = _key_value_block(ws, row, mb)
    row += 1

    # ---- Heat transfer ----
    row = _section_header(ws, row, '  HEAT TRANSFER & OPERATING CONDITIONS', span_cols=5)
    ht = [
        ('Total heat load',          f"{_fmt(atfd_r.get('Q_total_kcalh'), 0)} kcal/h"),
        ('Sensible heat load',       f"{_fmt(atfd_r.get('Q_sensible_kcalh'), 0)} kcal/h"),
        ('Latent heat load',         f"{_fmt(atfd_r.get('Q_latent_kcalh'), 0)} kcal/h"),
        ('HTA calculated',           f"{_fmt(atfd_r.get('HTA_calc_m2'), 2)} m²"),
        ('HTA selected',             f"{_fmt(atfd_r.get('HTA_selected_m2'), 0)} m²"),
        ('Overall U',                f"{_fmt(atfd_r.get('U_dryer'), 0)} kcal/h·m²·°C"),
        ('LMTD',                     f"{_fmt(atfd_r.get('lmtd_c'), 1)} °C"),
        ('Shell temperature',        f"{_fmt(atfd_r.get('shell_temp_c'), 1)} °C"),
        ('Shell pressure',           f"{_fmt(atfd_r.get('shell_pressure_bar'), 2)} bar-a"),
        ('Product boiling temp',     f"{_fmt(atfd_r.get('boiling_temp_c'), 1)} °C"),
        ('Boiling point elevation',  f"{_fmt(atfd_r.get('bpe_c'), 1)} °C"),
        ('Steam consumption',        f"{_fmt(atfd_r.get('steam_consumption_kgh'), 0)} kg/h"),
    ]
    row = _key_value_block(ws, row, ht)
    row += 1

    # ---- Equipment ----
    row = _section_header(ws, row, '  EQUIPMENT', span_cols=5)
    eq = [
        ('Agitator motor (selected)', f"{_fmt(atfd_r.get('motor_hp'), 0)} HP"),
        ('Connected load',           f"{_fmt(atfd_r.get('connected_load_kw'), 1)} kW"),
        ('Power consumed',           f"{_fmt(atfd_r.get('power_consumed_kwh'), 2)} kWh"),
    ]
    # Condenser
    acond = atfd_r.get('condenser') or {}
    if acond:
        eq.extend([
            ('Condenser HTA (calc)',     f"{_fmt(acond.get('HTA_calc_m2'), 1)} m²"),
            ('Condenser HTA (selected)', f"{_fmt(acond.get('HTA_selected_m2'), 0)} m²"),
            ('Condenser CW flow',        f"{_fmt(acond.get('cw_flow_m3h'), 1)} m³/h"),
        ])
    # Blower
    ablow = atfd_r.get('blower') or {}
    if ablow:
        eq.extend([
            ('Blower motor',             f"{_fmt(ablow.get('motor_hp'), 1)} HP"),
            ('Blower flow',              f"{_fmt(ablow.get('flow_m3h'), 0)} m³/h"),
        ])
    row = _key_value_block(ws, row, eq)
    row += 1

    # ---- Pumps ----
    pumps = atfd_r.get('pumps') or {}
    if pumps:
        row = _section_header(ws, row, '  ATFD PUMPS', span_cols=5)
        pump_rows = []
        for pid, p in pumps.items():
            pump_rows.append((
                p.get('service', pid),
                _fmt(p.get('flow_m3h'), 2) + ' m³/h',
                _fmt(p.get('head_mlc'), 0) + ' m',
                _fmt(p.get('brake_power_kw'), 2) + ' kW',
                _fmt(p.get('motor_hp_selected'), 1) + ' HP',
            ))
        _data_table(ws, row,
                    headers=['Service', 'Flow', 'Head', 'Brake Power', 'Motor'],
                    rows=pump_rows,
                    col_widths=[32, 16, 12, 16, 14])


def _build_plantwide_sheet(ws, pw, proj) -> None:
    _set_col_widths(ws, {'A': 32, 'B': 22, 'C': 18, 'D': 18, 'E': 18, 'F': 18})

    row = _title_bar(ws, 1, 'PLANT-WIDE SUMMARY',
                     'Consolidated utilities, pumps, feed trace, and economics',
                     span_cols=6)

    # ---- Total utilities ----
    u = pw.get('total_utilities') or {}
    if u:
        row = _section_header(ws, row, '  TOTAL UTILITIES', span_cols=6)
        ui = [
            ('Total steam demand',       f"{_fmt(u.get('steam_kgh'), 0)} kg/h   ({_fmt((u.get('steam_kgh') or 0) / 1000, 1)} TPH)"),
            ('Total power demand',       f"{_fmt(u.get('power_kw'), 1)} kW"),
            ('Total cooling water',      f"{_fmt(u.get('cw_m3h'), 0)} m³/h"),
        ]
        row = _key_value_block(ws, row, ui)
        row += 1

    # ---- Equipment summary ----
    ec = pw.get('equipment_count') or {}
    if ec:
        row = _section_header(ws, row, '  EQUIPMENT SUMMARY', span_cols=6)
        ei = [
            ('Stripper columns',         f"{ec.get('stripper_column') or '—'}"),
            ('MEE effects',              f"{ec.get('mee_effects') or '—'}"),
            ('MEE vapor-liquid separators', f"{ec.get('mee_vls') or '—'}"),
            ('MEE pre-heaters',          f"{ec.get('mee_preheaters') or '—'}"),
            ('Heat exchangers (total)',  f"{ec.get('heat_exchangers_total') or '—'}"),
            ('Pumps (total)',            f"{pw.get('total_pumps_count') or ec.get('pumps_total') or '—'}"),
            ('Total motor HP',           f"{_fmt(pw.get('total_motor_hp'), 1)} HP"),
        ]
        row = _key_value_block(ws, row, ei)
        row += 1

    # ---- Consolidated pump schedule ----
    pumps = pw.get('consolidated_pump_list') or []
    if pumps:
        row = _section_header(ws, row, '  CONSOLIDATED PUMP SCHEDULE', span_cols=6)
        pump_rows = []
        for p in pumps:
            pump_rows.append((
                p.get('unit', '—'),
                p.get('service', '—'),
                _fmt(p.get('flow_m3h'), 2) + ' m³/h',
                _fmt(p.get('head_mlc'), 0) + ' m',
                _fmt(p.get('brake_power_kw'), 2) + ' kW',
                _fmt(p.get('motor_hp_selected'), 1) + ' HP',
            ))
        row = _data_table(ws, row,
                          headers=['Unit', 'Service', 'Flow', 'Head', 'Brake Power', 'Motor'],
                          rows=pump_rows,
                          col_widths=[14, 32, 16, 12, 16, 14])
        row += 1

    # ---- Feed characterization trace ----
    trace = pw.get('feed_characterization_traceability') or []
    if trace:
        row = _section_header(ws, row, '  FEED CHARACTERIZATION TRACE', span_cols=6)
        params = [
            ('pH', 'ph', '', 2),
            ('Total Solids %', 'ts_pct', '%', 2),
            ('TDS %', 'tds_pct', '%', 2),
            ('TSS %', 'tss_pct', '%', 2),
            ('COD', 'cod_mgl', 'mg/L', 0),
            ('BOD', 'bod_mgl', 'mg/L', 0),
            ('Chlorides', 'chlorides_mgl', 'mg/L', 0),
            ('Sulphates', 'sulphates_mgl', 'mg/L', 0),
            ('Crystalline salt %', 'crystalline_salt_pct', '%', 1),
            ('Non-crystalline salt %', 'non_crystalline_salt_pct', '%', 1),
        ]
        stage_names = [t.get('stage', f'Stage {i+1}') for i, t in enumerate(trace[:4])]
        headers = ['Parameter'] + stage_names

        trace_rows = []
        for plabel, pkey, unit, places in params:
            vals = []
            for stage in trace[:4]:
                ch = stage.get('characterization') or {}
                v = ch.get(pkey)
                if v is None:
                    vals.append('—')
                else:
                    vals.append(_fmt(v, places))
            param_display = plabel if not unit else plabel
            trace_rows.append([param_display] + vals)

        col_widths = [26] + [18] * len(stage_names)
        row = _data_table(ws, row, headers=headers, rows=trace_rows, col_widths=col_widths)
        row += 1

    # ---- Salt routing ----
    sr = pw.get('salt_routing') or {}
    if sr:
        row = _section_header(ws, row, '  SALT ROUTING', span_cols=6)
        sr_items = [
            ('Total solids',             f"{_fmt(sr.get('total_solids_kgh'), 1)} kg/h"),
            ('Crystalline salt',         f"{_fmt(sr.get('crystalline_salt_kgh'), 1)} kg/h"),
            ('Non-crystalline salt',     f"{_fmt(sr.get('non_crystalline_salt_kgh'), 1)} kg/h"),
            ('Precipitated salt',        f"{_fmt(sr.get('precipitated_salt_kgh'), 1)} kg/h"),
            ('Remaining in ML',          f"{_fmt(sr.get('remaining_in_ml_kgh'), 1)} kg/h"),
            ('MEE outlet TS %',          f"{_fmt(sr.get('mee_outlet_ts_pct'), 1)} %"),
            ('Crystallization saturation', f"{_fmt(sr.get('crystallization_saturation_pct'), 1)} %"),
        ]
        row = _key_value_block(ws, row, sr_items)
        row += 1

    # ---- Economics ----
    e = pw.get('economics') or {}
    if e:
        row = _section_header(ws, row, '  OPERATING ECONOMICS', span_cols=6)
        econ = [
            ('Cost per KL treated',      f"₹ {_fmt(e.get('cost_per_kl_inr'), 0)}"),
            ('Daily operating cost',     f"₹ {_fmt(e.get('total_daily_op_cost_inr'), 0)}"),
            ('Annual operating cost',    f"₹ {_fmt(e.get('annual_op_cost_inr'), 0)}   ({_fmt((e.get('annual_op_cost_inr') or 0) / 10000000, 2)} crore)"),
            ('Operating days / year',    f"{e.get('operating_days_year', '—')}"),
            ('Operating hours / day',    f"{e.get('operating_hours_day', '—')}"),
            ('Daily steam cost',         f"₹ {_fmt(e.get('daily_steam_cost_inr'), 0)}"),
            ('Daily power cost',         f"₹ {_fmt(e.get('daily_power_cost_inr'), 0)}"),
            ('Daily cooling water cost', f"₹ {_fmt(e.get('daily_cw_cost_inr'), 0)}"),
        ]
        row = _key_value_block(ws, row, econ)
        row += 1


# =================================================================
# Public entry point
# =================================================================
def build_review_workbook(data: dict) -> bytes:
    """
    Build a 5-sheet branded Excel workbook for manager review.

    Args:
      data: project export dict (same structure as build_full_project_export).

    Returns:
      XLSX bytes suitable for st.download_button.
    """
    proj = data.get('project') or {}
    po = data.get('plant_overview') or {}
    strip_d = data.get('stripper') or {}
    mee_d = data.get('mee') or {}
    atfd_d = data.get('atfd') or {}
    pw = data.get('plant_wide') or {}
    e = pw.get('economics') or {}

    strip_i = strip_d.get('inputs') or {}
    strip_r = strip_d.get('results') or {}
    mee_i = mee_d.get('inputs') or {}
    mee_r = mee_d.get('results') or {}
    atfd_i = atfd_d.get('inputs') or {}
    atfd_r = atfd_d.get('results') or {}

    has_strip = bool(strip_r.get('feed_kgh'))
    has_mee = bool(mee_r.get('feed_kgh'))
    has_atfd = bool(atfd_r.get('feed_kgh'))

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = 'Project Summary'
    _build_summary_sheet(ws_summary, proj, po, pw, e, has_strip, has_mee, has_atfd)

    # Sheet 2: Stripper
    ws_strip = wb.create_sheet('Stripper')
    _build_stripper_sheet(ws_strip, strip_i, strip_r)

    # Sheet 3: MEE
    ws_mee = wb.create_sheet('MEE')
    _build_mee_sheet(ws_mee, mee_i, mee_r)

    # Sheet 4: ATFD
    ws_atfd = wb.create_sheet('ATFD')
    _build_atfd_sheet(ws_atfd, atfd_i, atfd_r)

    # Sheet 5: Plant-Wide
    ws_pw = wb.create_sheet('Plant-Wide')
    _build_plantwide_sheet(ws_pw, pw, proj)

    # Hide gridlines on all sheets for a cleaner look
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
